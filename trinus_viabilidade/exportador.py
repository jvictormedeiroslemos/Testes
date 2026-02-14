"""
Trinus Viabilidade - Exportação de premissas.

Exporta resultados em JSON e Excel.
"""

from __future__ import annotations

import json
from io import BytesIO
from datetime import datetime
from typing import TYPE_CHECKING

import pandas as pd

from modelos import ResultadoPremissas

if TYPE_CHECKING:
    from simulador import ResultadoSimulacao


def exportar_json(resultado: ResultadoPremissas) -> str:
    """Exporta premissas para JSON formatado."""
    data = resultado.to_dict()
    data["exportado_em"] = datetime.now().isoformat()
    return json.dumps(data, ensure_ascii=False, indent=2)


def exportar_json_bytes(resultado: ResultadoPremissas) -> bytes:
    """Exporta premissas para bytes JSON (para download)."""
    return exportar_json(resultado).encode("utf-8")


def _premissas_para_dataframe(resultado: ResultadoPremissas) -> pd.DataFrame:
    """Converte premissas para DataFrame do pandas."""
    rows = []
    for p in resultado.premissas:
        rows.append({
            "Categoria": p.categoria,
            "Subcategoria": p.subcategoria,
            "Premissa": p.nome,
            "Valor": p.valor,
            "Unidade": p.unidade,
            "Mínimo": p.valor_min,
            "Máximo": p.valor_max,
            "Fonte": p.fonte,
            "Descrição": p.descricao,
        })
    return pd.DataFrame(rows)


def _inputs_para_dataframe(resultado: ResultadoPremissas) -> pd.DataFrame:
    """Converte inputs do usuário para DataFrame."""
    inp = resultado.inputs
    rows = [
        {"Campo": "Tipologia", "Valor": inp.tipologia.value},
        {"Campo": "Estado", "Valor": inp.estado},
        {"Campo": "Cidade", "Valor": inp.cidade},
        {"Campo": "Região", "Valor": inp.regiao.value},
        {"Campo": "Padrão", "Valor": inp.padrao.value},
        {"Campo": "Nº Unidades", "Valor": inp.num_unidades},
        {"Campo": "Tipo de Negociação", "Valor": inp.tipo_negociacao.value},
        {"Campo": "Área do Terreno (m²)", "Valor": inp.area_terreno_m2 or "Não informado"},
        {"Campo": "VGV Estimado (R$)", "Valor": inp.vgv_estimado or "Calculado pelo sistema"},
        {"Campo": "Área Privativa Média (m²)", "Valor": inp.area_privativa_media_m2 or "Estimado pelo sistema"},
    ]
    # Parâmetros de negociação
    if inp.permuta_percentual is not None:
        rows.append({"Campo": "Permuta (%)", "Valor": inp.permuta_percentual})
    if inp.permuta_referencia:
        rows.append({"Campo": "Referência da Permuta", "Valor": inp.permuta_referencia})
    if inp.valor_aquisicao is not None:
        rows.append({"Campo": "Valor da Aquisição (R$)", "Valor": inp.valor_aquisicao})
    return pd.DataFrame(rows)


def _tabela_vendas_para_dataframe(resultado: ResultadoPremissas) -> pd.DataFrame:
    """Converte tabela de vendas para DataFrame (Loteamento ou Incorporação)."""
    # Loteamento
    if resultado.tabela_vendas_loteamento:
        tv = resultado.tabela_vendas_loteamento
        rows = [
            {"Condição": "Entrada", "Valor": f"{tv.entrada_pct}%"},
            {"Condição": "Nº parcelas da entrada", "Valor": tv.num_parcelas_entrada},
            {"Condição": "Saldo parcelado", "Valor": f"{tv.saldo_parcelado_pct}%"},
            {"Condição": "Nº parcelas", "Valor": tv.num_parcelas},
            {"Condição": "Sistema de amortização", "Valor": tv.sistema_amortizacao},
            {"Condição": "Juros (% a.m.)", "Valor": f"{tv.juros_am}%"},
            {"Condição": "Indexador", "Valor": tv.indexador},
            {"Condição": "Intermediárias", "Valor": f"{tv.intermediarias_pct}%"},
        ]
        return pd.DataFrame(rows)

    # Incorporação
    tv = resultado.tabela_vendas
    if not tv:
        return pd.DataFrame()
    rows = [
        {"Condição": "Entrada", "Valor": f"{tv.entrada_pct}%"},
        {"Condição": "Parcelas durante obra", "Valor": f"{tv.parcelas_obra_pct}%"},
        {"Condição": "Financiamento (chaves)", "Valor": f"{tv.financiamento_pct}%"},
        {"Condição": "Reforços", "Valor": f"{tv.reforcos_pct}%"},
        {"Condição": "Nº parcelas durante obra", "Valor": tv.num_parcelas_obra},
        {"Condição": "Indexador pré-chaves", "Valor": tv.indexador_pre_chaves},
        {"Condição": "Indexador pós-chaves", "Valor": tv.indexador_pos_chaves},
    ]
    return pd.DataFrame(rows)


def _resumo_dfc_dataframe(resultado: ResultadoPremissas) -> pd.DataFrame:
    """Gera resumo consolidado no formato DFC."""
    vgv_p = resultado.get_premissa("VGV estimado")
    vgv_val = vgv_p.valor if vgv_p else 0

    rows = []
    for cat in ["Receita", "Custo", "Despesa", "Financeiro"]:
        rows.append({"Categoria DFC": f"=== {cat.upper()} ===", "Premissa": "", "Valor": "", "Unidade": "", "Valor Estimado (R$)": ""})
        for p in resultado.por_categoria(cat):
            valor_abs = ""
            if p.unidade == "R$":
                valor_abs = f"R$ {p.valor:,.2f}"
            elif ("% do VGV" in p.unidade or "% sobre receita" in p.unidade) and vgv_val > 0:
                valor_abs = f"R$ {vgv_val * p.valor / 100:,.2f}"
            rows.append({
                "Categoria DFC": p.subcategoria,
                "Premissa": p.nome,
                "Valor": p.valor,
                "Unidade": p.unidade,
                "Valor Estimado (R$)": valor_abs,
            })
    return pd.DataFrame(rows)


def exportar_excel_bytes(resultado: ResultadoPremissas, sim: "ResultadoSimulacao | None" = None) -> bytes:
    """Exporta premissas para bytes Excel (para download)."""
    buffer = BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # Aba 1: Inputs
        df_inputs = _inputs_para_dataframe(resultado)
        df_inputs.to_excel(writer, index=False, sheet_name="Inputs")

        # Aba 2: Premissas completas
        df_premissas = _premissas_para_dataframe(resultado)
        df_premissas.to_excel(writer, index=False, sheet_name="Premissas")

        # Aba 3: Tabela de vendas
        df_vendas = _tabela_vendas_para_dataframe(resultado)
        if not df_vendas.empty:
            df_vendas.to_excel(writer, index=False, sheet_name="Tabela de Vendas")

        # Abas por categoria (dinâmico)
        categorias = df_premissas["Categoria"].unique()
        for cat in categorias:
            df_cat = df_premissas[df_premissas["Categoria"] == cat].copy()
            if not df_cat.empty:
                # Limitar nome da aba a 31 caracteres (limite Excel)
                sheet_name = cat[:31]
                df_cat.to_excel(writer, index=False, sheet_name=sheet_name)

        # Aba Resumo DFC
        df_resumo = _resumo_dfc_dataframe(resultado)
        if not df_resumo.empty:
            df_resumo.to_excel(writer, index=False, sheet_name="Resumo DFC")

        # Aba DFC Aberto (mês a mês) — se simulação disponível
        if sim is not None and sim.total_meses > 0:
            df_dfc = dfc_aberto_dataframe(sim, resultado.e_loteamento)
            if not df_dfc.empty:
                df_dfc.to_excel(writer, index=False, sheet_name="DFC")

    return buffer.getvalue()


def dfc_aberto_dataframe(sim: "ResultadoSimulacao", e_loteamento: bool = False) -> pd.DataFrame:
    """Gera DataFrame do DFC Aberto mês a mês (estrutura padrão Trinus).

    Colunas = itens financeiros hierárquicos, Linhas = meses.
    Inclui linhas de totais (Nominal) no topo.
    """
    n = sim.total_meses
    if n == 0:
        return pd.DataFrame()

    lbl_obra = "infraestrutura" if e_loteamento else "construção"
    lbl_inter = "Intermediárias / Saldo" if e_loteamento else "Financiamento / Reforços"

    # Helper: safe access to arrays (pad with 0 if shorter)
    def _arr(arr: list[float]) -> list[float]:
        if len(arr) >= n:
            return arr[:n]
        return arr + [0.0] * (n - len(arr))

    # Build ordered dict of columns (preserving insertion order = hierarchy)
    from collections import OrderedDict
    cols = OrderedDict()

    cols["Mês"] = list(range(n))

    # --- VENDAS ---
    cols["Vendas (unidades)"] = _arr(sim.vendas_unidades_mensal)
    cols["Vendas acumuladas (un.)"] = _arr(sim.vendas_unidades_acum_mensal)
    cols["VGV vendido"] = _arr(sim.vgv_vendido_mensal)

    # --- RECEITA ---
    cols["Receita Líquida Operacional"] = _arr(sim.receitas_mensais)
    cols["  Receita Bruta Operacional"] = _arr(sim.receita_bruta_mensal)
    cols["    Entrada / Sinal"] = _arr(sim.rec_entrada_mensal)
    cols["    Parcelas mensais"] = _arr(sim.rec_parcelas_mensal)
    cols[f"    {lbl_inter}"] = _arr(sim.rec_intermediarias_mensal)
    cols["  (-) Inadimplência"] = [-v for v in _arr(sim.inadimplencia_mensal)]

    # --- DEDUTORES DE RECEITA ---
    cols["(-) Impostos sobre receita"] = [-v for v in _arr(sim.impostos_mensal)]

    # --- CUSTOS ---
    custos_m = _arr(sim.custos_mensais)
    cols["(-) Custo total"] = [-v for v in custos_m]

    # Custo de obra
    obra_raso = _arr(sim.custo_obra_raso_mensal)
    admin_obra = _arr(sim.custo_admin_obra_mensal)
    projetos = _arr(sim.custo_projetos_mensal)
    aprovacoes = _arr(sim.custo_aprovacoes_mensal)
    obra_total = [obra_raso[i] + admin_obra[i] + projetos[i] + aprovacoes[i] for i in range(n)]
    cols["  Custo total de obra"] = [-v for v in obra_total]
    cols[f"    Custo raso de {lbl_obra}"] = [-v for v in obra_raso]
    cols["    Custo de administração (BDI)"] = [-v for v in admin_obra]
    cols["    Custo de projetos"] = [-v for v in projetos]
    cols["    Custo de aprovações"] = [-v for v in aprovacoes]

    # Custo de terreno
    terreno = _arr(sim.custo_terreno_aquisicao_mensal)
    itbi = _arr(sim.custo_itbi_mensal)
    iptu = _arr(sim.custo_iptu_mensal)
    terreno_total = [terreno[i] + itbi[i] + iptu[i] for i in range(n)]
    cols["  Custo de terreno"] = [-v for v in terreno_total]
    cols["    Aquisição do terreno"] = [-v for v in terreno]
    cols["    ITBI"] = [-v for v in itbi]
    cols["    IPTU"] = [-v for v in iptu]

    # --- DESPESAS ---
    despesas_m = _arr(sim.despesas_mensais)
    cols["(-) Total de despesas"] = [-v for v in despesas_m]

    # Despesas comerciais
    comissoes = _arr(sim.desp_comissoes_mensal)
    premiacao = _arr(sim.desp_premiacao_mensal)
    stand = _arr(sim.desp_stand_mensal)
    coordenacao = _arr(sim.desp_coordenacao_mensal)
    comercial = [comissoes[i] + premiacao[i] + stand[i] + coordenacao[i] for i in range(n)]
    cols["  Despesas comerciais"] = [-v for v in comercial]
    cols["    Comissões (corretagem)"] = [-v for v in comissoes]
    cols["    Premiação de corretores"] = [-v for v in premiacao]
    cols["    Central de vendas / Stand"] = [-v for v in stand]
    cols["    Coordenação comercial"] = [-v for v in coordenacao]

    # Marketing
    marketing = _arr(sim.desp_marketing_mensal)
    cols["  Despesas de marketing"] = [-v for v in marketing]

    # Despesas administrativas
    admin = _arr(sim.desp_admin_mensal)
    gestao = _arr(sim.desp_gestao_mensal)
    seguros = _arr(sim.desp_seguros_mensal)
    preop = _arr(sim.desp_preop_mensal)
    adm_total = [admin[i] + gestao[i] + seguros[i] + preop[i] for i in range(n)]
    cols["  Despesas administrativas"] = [-v for v in adm_total]
    cols["    Gerenciamento / Taxa de gestão"] = [-v for v in gestao]
    cols["    Administrativas SPE"] = [-v for v in admin]
    cols["    Seguros"] = [-v for v in seguros]
    cols["    Pré-operacionais"] = [-v for v in preop]

    # Despesas cartoriais
    ri = _arr(sim.desp_ri_mensal)
    escrituras = _arr(sim.desp_escrituras_mensal)
    cartorial = [ri[i] + escrituras[i] for i in range(n)]
    cols["  Despesas cartoriais"] = [-v for v in cartorial]
    cols["    Registro de incorporação"] = [-v for v in ri]
    cols["    Escrituras e registros"] = [-v for v in escrituras]

    # Tributária (já está em despesas, mas mostra separada)
    tributaria = _arr(sim.desp_tributaria_mensal)
    cols["  Impostos sobre receita"] = [-v for v in tributaria]

    # --- RESULTADO ---
    cols["(=) Atividades operacionais"] = _arr(sim.atividades_operacionais_mensal)
    cols["Saldo em caixa (acumulado)"] = _arr(sim.saldo_caixa_mensal)

    df = pd.DataFrame(cols)

    # Adicionar linhas de totais no topo (Nominal)
    nominal = {}
    for col in df.columns:
        if col == "Mês":
            nominal[col] = "NOMINAL"
        elif "acumulad" in col.lower() or "Saldo" in col:
            nominal[col] = df[col].iloc[-1] if len(df) > 0 else 0
        else:
            nominal[col] = df[col].sum()

    # Margem Nominal (% sobre Receita Líquida)
    rl_total = df["Receita Líquida Operacional"].sum()
    margem = {}
    for col in df.columns:
        if col == "Mês":
            margem[col] = "MARGEM %"
        elif rl_total != 0:
            margem[col] = f"{nominal[col] / rl_total * 100:.1f}%" if isinstance(nominal[col], (int, float)) else ""
        else:
            margem[col] = ""

    # % VGV
    vgv_val = sim.vgv
    pct_vgv = {}
    for col in df.columns:
        if col == "Mês":
            pct_vgv[col] = "% VGV"
        elif vgv_val != 0:
            pct_vgv[col] = f"{nominal[col] / vgv_val * 100:.1f}%" if isinstance(nominal[col], (int, float)) else ""
        else:
            pct_vgv[col] = ""

    # Prepend summary rows
    df_nominal = pd.DataFrame([nominal])
    df_margem = pd.DataFrame([margem])
    df_pct_vgv = pd.DataFrame([pct_vgv])
    df = pd.concat([df_nominal, df_margem, df_pct_vgv, df], ignore_index=True)

    return df


def exportar_dfc_aberto_excel_bytes(sim: "ResultadoSimulacao", e_loteamento: bool = False) -> bytes:
    """Exporta DFC Aberto mês a mês para Excel."""
    from simulador import ResultadoSimulacao

    buffer = BytesIO()
    df = dfc_aberto_dataframe(sim, e_loteamento)

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="DFC")

        # Formatação
        ws = writer.sheets["DFC"]
        from openpyxl.styles import Font, Alignment, numbers, PatternFill, Border, Side

        # Largura de colunas
        ws.column_dimensions["A"].width = 12  # Mês
        for col_idx in range(2, ws.max_column + 1):
            ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = 18

        # Estilo das linhas de cabeçalho (primeiras 4 linhas: header + 3 resumo)
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        summary_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
        summary_font = Font(bold=True, size=10)

        # Header row (row 1)
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(1, col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Summary rows (rows 2-4: Nominal, Margem, %VGV)
        for row in range(2, 5):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row, col)
                cell.fill = summary_fill
                cell.font = summary_font
                cell.alignment = Alignment(horizontal="center")

        # Formato numérico para dados mensais
        num_fmt = '#,##0'
        for row in range(5, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row, col)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = num_fmt
                    cell.alignment = Alignment(horizontal="right")

        # Destacar linhas de subtotais (colunas que não começam com espaço)
        subtotal_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        subtotal_font = Font(bold=True)
        # Identificar colunas de subtotal pelo header
        subtotal_cols = set()
        for col in range(2, ws.max_column + 1):
            header_val = str(ws.cell(1, col).value or "")
            if header_val and not header_val.startswith("  "):
                subtotal_cols.add(col)

        # Congelar painéis (fixar 4 linhas de cima + coluna A)
        ws.freeze_panes = "B5"

    return buffer.getvalue()


def premissas_para_dataframe(resultado: ResultadoPremissas) -> pd.DataFrame:
    """Interface pública para converter premissas em DataFrame."""
    return _premissas_para_dataframe(resultado)


def inputs_para_dataframe(resultado: ResultadoPremissas) -> pd.DataFrame:
    """Interface pública para converter inputs em DataFrame."""
    return _inputs_para_dataframe(resultado)


def tabela_vendas_para_dataframe(resultado: ResultadoPremissas) -> pd.DataFrame:
    """Interface pública para converter tabela de vendas em DataFrame."""
    return _tabela_vendas_para_dataframe(resultado)
